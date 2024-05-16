import os
from datetime import datetime

import flet as ft
import pandas as pd
from flet import View
from openpyxl import load_workbook
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, Boolean
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker

# SQLAlchemyの設定
db_url = "sqlite:///ldtp_app.db"
engine = create_engine(db_url)
# engine = create_engine("postgresql+psycopg2://postgres:postgres@localhost/fletapp")
# engine = create_engine("postgresql+psycopg2://postgres:postgres@192.168.3.5/fletapp")
Session = sessionmaker(bind=engine)
Base = declarative_base()

selected_row = None


# PatientInfoモデルの定義
class PatientInfo(Base):
    __tablename__ = 'patient_info'
    id = Column(Integer, primary_key=True)
    patient_id = Column(Integer)
    patient_name = Column(String)
    kana = Column(String)
    gender = Column(String)
    birthdate = Column(Date)
    issue_date = Column(Date)
    doctor_id = Column(Integer)
    doctor_name = Column(String)
    department = Column(String)
    main_diagnosis = Column(String)
    creation_count = Column(Integer)
    target_weight = Column(Float)
    sheet_name = Column(String)
    goal1 = Column(String)
    goal2 = Column(String)
    diet = Column(String)
    exercise_prescription = Column(String)
    exercise_time = Column(String)
    exercise_frequency = Column(String)
    exercise_intensity = Column(String)
    daily_activity = Column(String)
    nonsmoker = Column(String)
    smoking_cessation = Column(String)
    other1 = Column(String)
    other2 = Column(String)
    template = Column(String)


class MainDisease(Base):
    __tablename__ = "main_diseases"
    id = Column(Integer, primary_key=True)
    name = Column(String)


class SheetName(Base):
    __tablename__ = "sheet_names"
    id = Column(Integer, primary_key=True)
    main_disease_id = Column(Integer)
    name = Column(String)

class Template(Base):
    __tablename__ = 'templates'
    id = Column(Integer, primary_key=True)
    main_disease = Column(String)
    sheet_name = Column(String)
    goal1 = Column(String)
    goal2 = Column(String)
    diet = Column(String)
    exercise_prescription = Column(String)
    exercise_time = Column(String)
    exercise_frequency = Column(String)
    exercise_intensity = Column(String)
    daily_activity = Column(String)
    nonsmoker = Column(Boolean)
    other1 = Column(String)
    other2 = Column(String)

# テーブルの作成
Base.metadata.create_all(engine)


class TemplateEditor(ft.Control):
    def build(self):
        self.main_disease_dropdown = ft.Dropdown(label="主病名", options=load_main_diseases(), width=200)
        self.sheet_name_dropdown = ft.Dropdown(label="シート名", width=300)
        self.goal1 = ft.TextField(label="①達成目標：患者と相談した目標", width=700)
        self.goal2 = ft.TextField(label="②行動目標：患者と相談した目標", width=700)
        self.diet = ft.TextField(label="食事", multiline=True, width=400)
        self.exercise_prescription = ft.TextField(label="運動処方", width=200)
        self.exercise_time = ft.TextField(label="時間", width=200)
        self.exercise_frequency = ft.TextField(label="頻度", width=300)
        self.exercise_intensity = ft.TextField(label="強度", width=300)
        self.daily_activity = ft.TextField(label="日常生活の活動量増加", width=400)
        self.nonsmoker = ft.Checkbox(label="非喫煙者である")
        self.smoking_cessation = ft.Checkbox(label="禁煙の実施方法等を指示")
        self.other1 = ft.TextField(label="その他1", width=400)
        self.other2 = ft.TextField(label="その他2", width=400)

        self.save_button = ft.ElevatedButton("保存", on_click=self.save_template)
        self.cancel_button = ft.ElevatedButton("キャンセル", on_click=self.cancel_edit)

        return ft.Column([
            ft.Row([self.main_disease_dropdown, self.sheet_name_dropdown]),
            self.goal1,
            self.goal2,
            self.diet,
            ft.Row([self.exercise_prescription, self.exercise_time, self.exercise_frequency, self.exercise_intensity]),
            self.daily_activity,
            ft.Row([self.nonsmoker, self.smoking_cessation]),
            ft.Row([self.other1, self.other2]),
            ft.Row([self.save_button, self.cancel_button]),
        ])

    def save_template(self, e):
        # テンプレートを保存する処理を実装
        pass

    def cancel_edit(self, e):
        # テンプレート編集をキャンセルする処理を実装
        pass


class TemplateManager:
    def __init__(self):
        self.templates = {}

    def add_template(self, main_disease, sheet_name, template_data):
        self.templates[(main_disease, sheet_name)] = template_data

    def get_template(self, main_disease, sheet_name):
        return self.templates.get((main_disease, sheet_name))


def load_patient_data():
    date_columns = [0, 6]
    return pd.read_csv(r"C:\InnoKarte\pat.csv", encoding="shift_jis", header=None, parse_dates=date_columns)


def load_main_diseases():
    session = Session()
    main_diseases = session.query(MainDisease).all()
    session.close()
    return [ft.dropdown.Option(str(disease.name)) for disease in main_diseases]


def load_sheet_names(main_disease):
    session = Session()
    if main_disease:
        sheet_names = session.query(SheetName).filter(SheetName.main_disease_id == main_disease).all()
    else:
        sheet_names = session.query(SheetName).all()
    session.close()
    return [ft.dropdown.Option(sheet.name) for sheet in sheet_names]


def format_date(date_str):
    if pd.isna(date_str):
        return ""
    return pd.to_datetime(date_str).strftime("%Y/%m/%d")


def main(page: ft.Page):
    page.title = "生活習慣病療養計画書"
    page.window_width = 1200
    page.window_height = 800

    # 初期データの挿入
    session = Session()
    if session.query(MainDisease).count() == 0:
        main_diseases = [
            MainDisease(id=1, name="高血圧"),
            MainDisease(id=2, name="脂質異常症"),
            MainDisease(id=3, name="糖尿病")
        ]
        session.add_all(main_diseases)
        session.commit()

    if session.query(SheetName).count() == 0:
        sheet_names = [
            SheetName(main_disease_id=1, name="血圧140-90以下"),
            SheetName(main_disease_id=1, name="血圧130-80以下"),
            SheetName(main_disease_id=2, name="LDL120以下"),
            SheetName(main_disease_id=2, name="LDL100以下"),
            SheetName(main_disease_id=2, name="LDL70以下"),
            SheetName(main_disease_id=3, name="HbAc８％"),
            SheetName(main_disease_id=3, name="HbAc７％"),
            SheetName(main_disease_id=3, name="HbAc６％"),
        ]
        session.add_all(sheet_names)
        session.commit()

    if session.query(Template).count() == 0:
        templates = [
            Template(main_disease="糖尿病", sheet_name="HbAc８％", goal1="HbA1ｃを低血糖に注意して下げる",
                     goal2="ストレッチを中心とした運動/間食の制限/糖質の制限",
                     diet="・食事量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ストレッチ運動", exercise_time="10分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="息切れしない程度", daily_activity="ストレッチ運動を主に行う", nonsmoker=True,
                     other1="睡眠の確保１日７時間", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="糖尿病", sheet_name="HbAc７％", goal1="HbA1ｃ７％を目標/体重を当初の－３Kgとする",
                     goal2="１日８０００歩以上の歩行/間食の制限/糖質の制限",
                     diet="・食事量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="ほぼ毎日",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日8000歩以上", nonsmoker=True,
                     other1="睡眠の確保１日７時間", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="糖尿病", sheet_name="HbAc６％", goal1="HbA1ｃを正常化",
                     goal2="１日５０００歩以上の歩行/間食の制限/糖質の制限",
                     diet="・食事量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="１週間に５回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上", nonsmoker=True,
                     other1="睡眠の確保１日７時間", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="高血圧", sheet_name="血圧130-80以下",
                     goal1="家庭血圧が測定でき、朝と就寝前のいずれかで130/80mmHg以下",
                     goal2="塩分を控えた食事と運動習慣を目標にする",
                     diet="・塩分量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上", nonsmoker=True,
                     other1="睡眠の確保１日７時間", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="高血圧", sheet_name="血圧140-90以下",
                     goal1="家庭血圧が測定でき、朝と就寝前のいずれかで140/90mmHg以下",
                     goal2="塩分を控えた食事と運動習慣を目標にする",
                     diet="・塩分量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ストレッチ運動", exercise_time="30分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="ストレッチ運動を主に行う", nonsmoker=True,
                     other1="睡眠の確保１日７時間", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="LDL120以下", goal1="LDLコレステロール＜120/TG＜150/HDL≧40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="・食事摂取量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上", nonsmoker=True,
                     other1="飲酒の制限、肥満度の改善", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="LDL100以下", goal1="LDLコレステロール＜100/TG＜150/HDL≧40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="・食事摂取量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上", nonsmoker=True,
                     other1="飲酒の制限、肥満度の改善", other2="家庭での毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="LDL70以下", goal1="LDLコレステロール＜100/TG＜150/HDL≧40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="・脂肪の多い食品や甘い物を控える\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上", exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上", nonsmoker=True,
                     other1="飲酒の制限、肥満度の改善", other2="家庭での毎日の歩数の測定"),
        ]
        session.add_all(templates)
        session.commit()

    session.close()

    def on_main_diagnosis_change(e):
        selected_main_disease = main_diagnosis.value
        apply_template()
        if selected_main_disease:
            session = Session()
            main_disease = session.query(MainDisease).filter(MainDisease.name == selected_main_disease).first()
            session.close()
            if main_disease:
                sheet_name_options = load_sheet_names(main_disease.id)
            else:
                sheet_name_options = []
        else:
            sheet_name_options = load_sheet_names(None)

        sheet_name_dropdown.options = sheet_name_options
        sheet_name_dropdown.value = ""
        page.update()

    def on_sheet_name_change(e):
        apply_template()
        page.update()

    df_patients = load_patient_data()
    print(df_patients)
    # CSVファイルから1行目の患者IDを取得
    initial_patient_id = ""
    if not df_patients.empty:
        initial_patient_id = str(df_patients.iloc[0, 2])

    def load_patient_info(patient_id):
        patient_info = df_patients[df_patients.iloc[:, 2] == patient_id]
        if not patient_info.empty:
            patient_info = patient_info.iloc[0]
            issue_date_value.value = format_date(patient_info.iloc[0])
            name_value.value = patient_info.iloc[3]
            kana_value.value = patient_info.iloc[4]
            gender_value.value = "男性" if patient_info.iloc[5] == 1 else "女性"
            birthdate = patient_info.iloc[6]
            birthdate_value.value = format_date(birthdate)
            doctor_id_value.value = str(patient_info.iloc[9])
            doctor_name_value.value = patient_info.iloc[10]
            department_value.value = patient_info.iloc[14]
        else:
            # patient_infoが空の場合は空文字列を設定
            issue_date_value.value = ""
            name_value.value = ""
            kana_value.value = ""
            gender_value.value = ""
            birthdate_value.value = ""
            doctor_id_value.value = ""
            doctor_name_value.value = ""
            department_value.value = ""
        page.update()

    def populate_common_sheet(common_sheet, patient_info):
        common_sheet["B2"] = patient_info.patient_id
        common_sheet["B3"] = patient_info.patient_name
        common_sheet["B4"] = patient_info.kana
        common_sheet["B5"] = patient_info.gender
        common_sheet["B6"] = patient_info.birthdate
        common_sheet["B7"] = patient_info.issue_date.strftime("%Y/%m/%d")
        common_sheet["B8"] = patient_info.doctor_id
        common_sheet["B9"] = patient_info.doctor_name
        common_sheet["B10"] = patient_info.department
        common_sheet["B11"] = patient_info.main_diagnosis
        common_sheet["B12"] = patient_info.creation_count
        common_sheet["B13"] = patient_info.target_weight
        common_sheet["B14"] = patient_info.sheet_name
        common_sheet["B15"] = patient_info.goal1
        common_sheet["B16"] = patient_info.goal2
        common_sheet["B17"] = patient_info.diet
        common_sheet["B18"] = patient_info.exercise_prescription
        common_sheet["B19"] = patient_info.exercise_time
        common_sheet["B20"] = patient_info.exercise_frequency
        common_sheet["B21"] = patient_info.exercise_intensity
        common_sheet["B22"] = patient_info.daily_activity
        common_sheet["B23"] = str(patient_info.nonsmoker)
        common_sheet["B24"] = str(patient_info.smoking_cessation)
        common_sheet["B25"] = patient_info.other1
        common_sheet["B26"] = patient_info.other2

    def create_treatment_plan(patient_id, doctor_id, doctor_name, department, df_patients):
        session = Session()
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        workbook = load_workbook(r"C:\Shinseikai\LDTPapp\生活習慣病療養計画書.xlsm", keep_vba=True)
        common_sheet = workbook["共通情報"]

        patient_info_csv = df_patients.loc[df_patients.iloc[:, 2] == patient_id]

        if patient_info_csv.empty:
            session.close()
            raise ValueError(f"患者ID {patient_id} が見つかりません。")

        patient_info = patient_info_csv.iloc[0]

        # データベースに保存
        treatment_plan = PatientInfo(
            patient_id=patient_id,
            patient_name=patient_info.iloc[3],
            kana=patient_info.iloc[4],
            gender="男性" if patient_info.iloc[5] == 1 else "女性",
            birthdate=patient_info.iloc[6],
            issue_date=datetime.now().date(),
            doctor_id=doctor_id,
            doctor_name=doctor_name,
            department=department,
            main_diagnosis=main_diagnosis.value,
            sheet_name=sheet_name_dropdown.value,
            creation_count=creation_count.value,
            target_weight=float(target_weight.value) if target_weight.value else None,
            goal1=goal1.value,
            goal2=goal2.value,
            diet=diet.value,
            exercise_prescription=exercise_prescription.value,
            exercise_time=exercise_time.value,
            exercise_frequency=exercise_frequency.value,
            exercise_intensity=exercise_intensity.value,
            daily_activity=daily_activity.value,
            nonsmoker=str(nonsmoker.value),
            smoking_cessation=str(smoking_cessation.value),
            other1=other1.value,
            other2=other2.value
        )
        session.add(treatment_plan)
        session.commit()

        populate_common_sheet(common_sheet, treatment_plan)

        new_file_name = f"生活習慣病療養計画書_{current_datetime}.xlsm"
        file_path = r"C:\Shinseikai\LDTPapp" + "\\" + new_file_name
        workbook.save(file_path)
        wb = load_workbook(file_path, read_only=False, keep_vba=True)
        ws_common = wb["共通情報"]
        ws_common.sheet_view.tabSelected = False
        ws_plan = wb["計画書"]
        ws_plan.sheet_view.tabSelected = True
        wb.active = ws_plan
        wb.save(file_path)
        os.startfile(file_path)

        session.close()
        open_route(None)

    def print_plan(e):
        global selected_row
        session = Session()
        if selected_row is not None:
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
                workbook = load_workbook(r"C:\Shinseikai\LDTPapp\生活習慣病療養計画書.xlsm", keep_vba=True)
                common_sheet = workbook["共通情報"]

                populate_common_sheet(common_sheet, patient_info)

                new_file_name = f"生活習慣病療養計画書_{current_datetime}.xlsm"
                file_path = r"C:\Shinseikai\LDTPapp" + "\\" + new_file_name
                workbook.save(file_path)
                wb = load_workbook(file_path, read_only=False, keep_vba=True)
                ws_common = wb["共通情報"]
                ws_common.sheet_view.tabSelected = False
                ws_plan = wb["計画書"]
                ws_plan.sheet_view.tabSelected = True
                wb.active = ws_plan
                wb.save(file_path)
                os.startfile(file_path)

        session.close()

    def create_new_plan(e):
        patient_id = patient_id_value.value.strip()
        doctor_id = doctor_id_value.value.strip()
        doctor_name = doctor_name_value.value
        if not patient_id or not doctor_id:
            page.snack_bar = ft.SnackBar(content=ft.Text("患者IDと医師IDは必須です"))
            page.snack_bar.open = True
            page.update()
            return
        department = department_value.value

        create_treatment_plan(int(patient_id), int(doctor_id), doctor_name, department, df_patients)

    def on_patient_id_change(e):
        patient_id = patient_id_value.value.strip()
        if patient_id:
            load_patient_info(int(patient_id))
        update_history(patient_id)

    def save_data(e):
        global selected_row
        session = Session()
        if selected_row is not None:
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                patient_info.patient_id = int(patient_id.value)
                patient_info.patient_name = name_value.value
                patient_info.kana = kana_value.value
                patient_info.gender = gender_value.value
                patient_info.birthdate = datetime.strptime(birthdate_value.value, "%Y/%m/%d").date()
                patient_info.issue_date = datetime.strptime(issue_date_value.value, "%Y/%m/%d").date()
                patient_info.doctor_id = int(doctor_id_value.value)
                patient_info.doctor_name = doctor_name_value.value
                patient_info.department = department_value.value
                patient_info.main_diagnosis = main_diagnosis.value
                patient_info.sheet_name = sheet_name_dropdown.value
                patient_info.creation_count = creation_count.value
                patient_info.target_weight = target_weight.value if target_weight.value else None
                patient_info.goal1 = goal1.value
                patient_info.goal2 = goal2.value
                patient_info.diet = diet.value
                patient_info.exercise_prescription = exercise_prescription.value
                patient_info.exercise_time = exercise_time.value
                patient_info.exercise_frequency = exercise_frequency.value
                patient_info.exercise_intensity = exercise_intensity.value
                patient_info.daily_activity = daily_activity.value
                patient_info.nonsmoker = str(nonsmoker.value)
                patient_info.smoking_cessation = str(smoking_cessation.value)
                patient_info.other1 = other1.value
                patient_info.other2 = other2.value
                session.commit()
                page.snack_bar = ft.SnackBar(
                    ft.Text("データが更新されました"),
                    duration=2000,
                )
                page.snack_bar.open = True
        else:
            patient_info = PatientInfo(
                patient_id=patient_id.value,
                patient_name=name_value.value,
                kana=kana_value.value,
                gender=gender_value.value,
                birthdate=datetime.strptime(birthdate_value.value, "%Y/%m/%d").date(),
                issue_date=datetime.now().date(),  # 現在の日付を設定
                doctor_id=int(doctor_id_value.value),  # doctor_id_value.valueを整数に変換して設定
                doctor_name=doctor_name_value.value,
                department=department_value.value,
                main_diagnosis=main_diagnosis.value,
                sheet_name=sheet_name_dropdown.value if sheet_name_dropdown.value else None,
                creation_count=creation_count.value,
                target_weight=target_weight.value if target_weight.value else None,
                goal1=goal1.value,
                goal2=goal2.value,
                diet=diet.value,
                exercise_prescription=exercise_prescription.value,
                exercise_time=exercise_time.value,
                exercise_frequency=exercise_frequency.value,
                exercise_intensity=exercise_intensity.value,
                daily_activity=daily_activity.value,
                nonsmoker=str(nonsmoker.value),
                smoking_cessation=str(smoking_cessation.value),
                other1=other1.value,
                other2=other2.value
            )
            session.add(patient_info)
            session.commit()
            page.snack_bar = ft.SnackBar(
                ft.Text("データが保存されました"),
                duration=2000,
            )
            page.snack_bar.open = True

        session.close()
        page.update()

    def copy_data(e):
        session = Session()
        patient_info = session.query(PatientInfo). \
            filter(PatientInfo.patient_id == patient_id.value). \
            order_by(PatientInfo.id.desc()).first()
        if patient_info:
            new_patient_info = PatientInfo(
                patient_id=patient_info.patient_id,
                patient_name=patient_info.patient_name,
                kana=patient_info.kana,
                gender=patient_info.gender,
                birthdate=patient_info.birthdate,
                issue_date=datetime.now().date(),
                doctor_id=patient_info.doctor_id,
                doctor_name=patient_info.doctor_name,
                department=patient_info.department,
                main_diagnosis=patient_info.main_diagnosis,
                sheet_name=patient_info.sheet_name,
                creation_count=patient_info.creation_count + 1,
                target_weight=patient_info.target_weight,
                goal1=patient_info.goal1,
                goal2=patient_info.goal2,
                diet=patient_info.diet,
                exercise_prescription=patient_info.exercise_prescription,
                exercise_time=patient_info.exercise_time,
                exercise_frequency=patient_info.exercise_frequency,
                exercise_intensity=patient_info.exercise_intensity,
                daily_activity=patient_info.daily_activity,
                nonsmoker=patient_info.nonsmoker,
                smoking_cessation=patient_info.smoking_cessation,
                other1=patient_info.other1,
                other2=patient_info.other2
            )
            session.add(new_patient_info)
            session.commit()
            selected_row = {"id": PatientInfo.id + 1}
            session.close()
            page.snack_bar = ft.SnackBar(
                ft.Text("前回データをコピーしました"),
                duration=2000,
            )
            page.snack_bar.open = True

        session.close()
        update_history(int(patient_id.value))
        page.update()

    def delete_data(e):
        session = Session()
        patient_info = session.query(PatientInfo).order_by(PatientInfo.id.desc()).first()
        if patient_info:
            session.delete(patient_info)
            session.commit()
            page.snack_bar = ft.SnackBar(
                ft.Text("データが削除されました"),
                duration=2000,
            )
            page.snack_bar.open = True
        session.close()
        open_route(e)

    def filter_data(e):
        update_history(patient_id.value)

    def update_history(filter_patient_id=None):
        data = fetch_data(filter_patient_id)
        history.rows = create_data_rows(data)
        page.update()

    def on_row_selected(e):
        global selected_row
        if e.data == "true":
            row_index = history.rows.index(e.control)
            selected_row = history.rows[row_index].data
            session = Session()
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                patient_id.value = patient_info.patient_id
                main_diagnosis.value = patient_info.main_diagnosis
                sheet_name_dropdown.value = patient_info.sheet_name
                creation_count.value = patient_info.creation_count
                target_weight.value = patient_info.target_weight
                goal1.value = patient_info.goal1
                goal2.value = patient_info.goal2
                diet.value = patient_info.diet
                exercise_prescription.value = patient_info.exercise_prescription
                exercise_time.value = patient_info.exercise_time
                exercise_frequency.value = patient_info.exercise_frequency
                exercise_intensity.value = patient_info.exercise_intensity
                daily_activity.value = patient_info.daily_activity
                nonsmoker.value = patient_info.nonsmoker == 'True'
                smoking_cessation.value = patient_info.smoking_cessation == 'True'
                other1.value = patient_info.other1
                other2.value = patient_info.other2
            session.close()
            page.update()  # 画面を更新

        if e.data == "true":
            row_index = history.rows.index(e.control)
            selected_row = history.rows[row_index].data
            open_edit(e)

    def fetch_data(filter_patient_id=None):
        if not filter_patient_id:
            return []

        session = Session()
        query = session.query(PatientInfo.id, PatientInfo.issue_date, PatientInfo.department,
                              PatientInfo.doctor_name, PatientInfo.main_diagnosis,
                              PatientInfo.sheet_name, PatientInfo.creation_count). \
            order_by(PatientInfo.patient_id.asc(), PatientInfo.id.desc())

        query = query.filter(PatientInfo.patient_id == filter_patient_id)

        patient_info_list = query.all()
        session.close()

        data = []
        for info in patient_info_list:
            data.append({
                "id": str(info.id),
                "issue_date": info.issue_date.strftime("%Y/%m/%d") if info.issue_date else "",
                "department": info.department,
                "doctor_name": info.doctor_name,
                "main_diagnosis": info.main_diagnosis,
                "sheet_name": info.sheet_name,
                "count": info.creation_count
            })

        return data

    def create_data_rows(data):
        rows = []
        for item in data:
            row = ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(item["id"])),
                    ft.DataCell(ft.Text(item["issue_date"])),
                    ft.DataCell(ft.Text(item["department"])),
                    ft.DataCell(ft.Text(item["doctor_name"])),
                    ft.DataCell(ft.Text(item["main_diagnosis"])),
                    ft.DataCell(ft.Text(item["sheet_name"])),
                    ft.DataCell(ft.Text(item["count"])),
                ],
                on_select_changed=on_row_selected,
                data=item
            )
            rows.append(row)
        return rows

    def apply_template(e=None):
        session = Session()
        template = session.query(Template).filter(Template.main_disease == main_diagnosis.value,
                                                  Template.sheet_name == sheet_name_dropdown.value).first()
        if template:
            goal1.value = template.goal1
            goal2.value = template.goal2
            diet.value = template.diet
            exercise_prescription.value = template.exercise_prescription
            exercise_time.value = template.exercise_time
            exercise_frequency.value = template.exercise_frequency
            exercise_intensity.value = template.exercise_intensity
            daily_activity.value = template.daily_activity
            nonsmoker.value = template.nonsmoker
            other1.value = template.other1
            other2.value = template.other2
        session.close()

    template_manager = TemplateManager()

    template_manager.add_template("糖尿病", "HbAc７％", {
        "goal1": "HbA1ｃ７％を目標/体重を当初の－３Kgとする",
        "goal2": "１日８０００歩以上の歩行/間食の制限/糖質の制限",
        "diet": "・食事量を適正にする\n・食物繊維の摂取量を増やす\n・ゆっくり食べる\n・間食を減らす",
        "exercise_prescription": "ウォーキング",
        "exercise_time": "30分以上",
        "exercise_frequency": "ほぼ毎日",
        "exercise_intensity": "少し汗をかく程度",
        "daily_activity": "1日8000歩以上",
        "nonsmoker": True,
        "other1": "睡眠の確保１日７時間",
        "other2": "家庭での毎日の歩数の測定",
    })

    def save_template(e):
        session = Session()
        template = session.query(Template).filter(Template.main_disease == main_diagnosis.value,
                                                  Template.sheet_name == sheet_name_dropdown.value).first()
        if template:
            template.goal1 = goal1.value
            template.goal2 = goal2.value
            template.diet = diet.value
            template.exercise_prescription = exercise_prescription.value
            template.exercise_time = exercise_time.value
            template.exercise_frequency = exercise_frequency.value
            template.exercise_intensity = exercise_intensity.value
            template.daily_activity = daily_activity.value
            template.nonsmoker = nonsmoker.value
            template.other1 = other1.value
            template.other2 = other2.value
        else:
            template = Template(
                main_disease=main_diagnosis.value,
                sheet_name=sheet_name_dropdown.value,
                goal1=goal1.value,
                goal2=goal2.value,
                diet=diet.value,
                exercise_prescription=exercise_prescription.value,
                exercise_time=exercise_time.value,
                exercise_frequency=exercise_frequency.value,
                exercise_intensity=exercise_intensity.value,
                daily_activity=daily_activity.value,
                nonsmoker=nonsmoker.value,
                other1=other1.value,
                other2=other2.value
            )
            session.add(template)
        session.commit()
        session.close()
        page.snack_bar = ft.SnackBar(
            content=ft.Text("テンプレートが保存されました"),
            duration=2000)
        page.snack_bar.open = True
        page.update()
        open_route(None)

    def route_change(e):
        print("Route change:", e.route)
        page.views.clear()
        page.views.append(
            View(
                "/",
                [
                    ft.Row(
                        controls=[
                            patient_id_value,
                            issue_date_value,
                            name_value,
                            kana_value,
                            gender_value,
                            birthdate_value
                        ]
                    ),
                    ft.Row(
                        controls=[
                            doctor_id_value,
                            doctor_name_value,
                            department_value,
                        ]
                    ),
                    ft.Row(
                        controls=[
                            buttons,
                        ]
                    ),
                    ft.Row(
                        controls=[
                            ft.Text("計画書一覧", size=16),
                            ft.Text("計画書をクリックすると編集画面が表示されます", size=14),
                        ]
                    ),
                    ft.Divider(),
                    history,
                ],
            )
        )

        if page.route == "/create":
            page.views.append(
                View(
                    "/create",
                    [
                        ft.Row(
                            controls=[
                                ft.Text("新規作成", size=14),
                                main_diagnosis,
                                sheet_name_dropdown,
                                creation_count,
                                ft.Text("回目", size=14),
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                                target_weight,
                                ft.Text("kg", size=14),
                            ]
                        ),
                        goal2,
                        guidance_items,
                        create_buttons,
                    ],
                )
            )
        page.update()

        if page.route == "/edit":
            page.views.append(
                View(
                    "/edit",
                    [
                        ft.Row(
                            controls=[
                                ft.Text("編集", size=14),
                                main_diagnosis,
                                sheet_name_dropdown,
                                creation_count,
                                ft.Text("回目", size=14),
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                                target_weight,
                                ft.Text("kg", size=14),
                            ]
                        ),
                        goal2,
                        guidance_items,
                        edit_buttons,
                    ],
                )
            )
        page.update()

        if page.route == "/templete":
            page.views.append(
                View(
                    "/templete",
                    [
                        ft.Row(
                            controls=[
                                ft.Text("テンプレート", size=14),
                                main_diagnosis,
                                sheet_name_dropdown,
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                            ]
                        ),
                        goal2,
                        guidance_items,
                        templete_buttons,
                    ],
                )
            )
        page.update()

    # 現在のページを削除して、前のページに戻る
    def view_pop(e):
        print("View pop:", e.view)
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    def open_create(e):
        page.go("/create")

    def open_edit(e):
        page.go("/edit")

    def open_templete(e):
        page.go("/templete")

    def open_route(e):
        for field in [main_diagnosis, target_weight, goal1, goal2, diet,
                      exercise_prescription, exercise_time, exercise_frequency, exercise_intensity,
                      daily_activity, other1, other2]:
            field.value = ""

        sheet_name_dropdown.value = ""
        creation_count.value = ""
        nonsmoker.value = False
        smoking_cessation.value = False

        page.go("/")
        update_history(int(patient_id.value))
        page.update()

    # Patient Information
    patient_id_value = ft.TextField(label="患者ID", on_change=on_patient_id_change, value=initial_patient_id, width=150)
    patient_id = ft.TextField(label="カルテID", width=150, on_change=on_patient_id_change,
                              value=initial_patient_id)  # patient_id_valueと区別するために変数名を変更
    issue_date_value = ft.TextField(label="発行日", read_only=True, width=150)
    name_value = ft.TextField(label="氏名", read_only=True, width=150)
    kana_value = ft.TextField(label="カナ", read_only=True, width=150)
    gender_value = ft.TextField(label="性別", read_only=True, width=150)
    birthdate_value = ft.TextField(label="生年月日", read_only=True, width=150)
    doctor_id_value = ft.TextField(label="医師ID", read_only=True, width=150)
    doctor_name_value = ft.TextField(label="医師名", read_only=True, width=150)
    department_value = ft.TextField(label="診療科", read_only=True, width=150)

    main_disease_options = load_main_diseases()
    main_diagnosis = ft.Dropdown(label="主病名", options=main_disease_options, width=200, value="",
                                 on_change=on_main_diagnosis_change, autofocus=True)
    sheet_name_options = load_sheet_names(main_diagnosis.value)
    sheet_name_dropdown = ft.Dropdown(label="シート名", options=sheet_name_options, width=300,value="",
                                      on_change=on_sheet_name_change, autofocus=True)
    creation_count = ft.TextField(
        label="作成回数",
        width=100,
        value="1",
        on_submit=lambda _: goal1.focus()
    )
    target_weight = ft.TextField(label="目標体重", width=100, value="", on_submit=lambda _: goal2.focus())
    goal1 = ft.TextField(label="①達成目標：患者と相談した目標", width=700, value="達成目標を入力してください",
                         on_submit=lambda _: target_weight.focus())
    goal2 = ft.TextField(label="②行動目標：患者と相談した目標", width=700, value="行動目標を入力してください",
                         on_submit=lambda _: diet.focus())

    diet = ft.TextField(
        label="食事",
        multiline=True,
        disabled=False,
        value="",
        width=400,
        on_submit=lambda _: exercise_prescription.focus()
    )

    exercise_prescription = ft.TextField(label="運動処方", width=200, value="ウォーキング",
                                         on_submit=lambda _: exercise_time.focus())
    exercise_time = ft.TextField(label="時間", value="30分以上", width=200,
                                 on_submit=lambda _: exercise_frequency.focus())
    exercise_frequency = ft.TextField(label="頻度", value="ほぼ毎日", width=300,
                                      on_submit=lambda _: exercise_intensity.focus())
    exercise_intensity = ft.TextField(label="強度", value="少し汗をかく程度", width=300,
                                      on_submit=lambda _: daily_activity.focus())
    daily_activity = ft.TextField(label="日常生活の活動量増加", value="1日8000歩以上", width=400,
                                  on_submit=lambda _: other1.focus())
    nonsmoker = ft.Checkbox(label="非喫煙者である", value=True)
    smoking_cessation = ft.Checkbox(label="禁煙の実施方法等を指示")
    other1 = ft.TextField(label="その他1", value="睡眠の確保１日７時間", width=400, on_submit=lambda _: other2.focus())
    other2 = ft.TextField(label="その他2", value="家庭での毎日の歩数の測定", width=400)

    guidance_items = ft.Column([
        diet,
        ft.Row([exercise_prescription, exercise_time, exercise_frequency, exercise_intensity]),
        daily_activity,
        ft.Row([ft.Text("たばこ", size=14), nonsmoker, smoking_cessation]),
        ft.Row([other1, other2]),
    ])

    selected_row = None
    data = fetch_data()
    rows = create_data_rows(data)

    history = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("ID")),
            ft.DataColumn(ft.Text("発行日")),
            ft.DataColumn(ft.Text("診療科")),
            ft.DataColumn(ft.Text("医師名")),
            ft.DataColumn(ft.Text("主病名")),
            ft.DataColumn(ft.Text("シート名")),
            ft.DataColumn(ft.Text("作成回数")),
        ],
        rows=rows,
        width=1200,
    )

    buttons = ft.Row([
        ft.ElevatedButton("新規作成", on_click=open_create),
        ft.ElevatedButton("前回コピー", on_click=copy_data),
        ft.ElevatedButton("テンプレート編集", on_click=open_templete),
    ])

    create_buttons = ft.Row([
        ft.ElevatedButton("新規発行", on_click=create_new_plan),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    edit_buttons = ft.Row([
        ft.ElevatedButton("保存", on_click=save_data),
        ft.ElevatedButton("印刷", on_click=print_plan),
        ft.ElevatedButton("削除", on_click=delete_data),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    templete_buttons = ft.Row([
        ft.ElevatedButton("保存", on_click=save_template),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    # page画面の設定
    layout = ft.Column([
        ft.Row(
            controls=[]
        ),
    ])

    page.add(layout)
    update_history()

    if initial_patient_id:
        load_patient_info(int(initial_patient_id))
        patient_id.value = initial_patient_id
        filter_data(patient_id.value)
        update_history(patient_id.value)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)


ft.app(target=main)
